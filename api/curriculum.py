from openpyxl import load_workbook  # openpyxl只支持读取xlsx的excel表格
from datetime import date, timedelta  # 获取日期
import re  # 正则表达式，匹配字符串


def send_class_table(raw_message):
    week_list = ['星期日', '星期一', '星期二', '星期三', '星期四', '星期五', '星期六']
    first_date = date(2023, 2, 13)  # 新学期第一周的第一天
    try:
        which_day = raw_message.split(' ')[1]
        if which_day == "今天":
            search_date = date.today()  # 获取今天的日期
        elif which_day == "明天":
            search_date = (date.today() + timedelta(days=1))  # 获取明天的日期
        elif which_day == "后天":
            search_date = (date.today() + timedelta(days=2))  # 获取后天的日期
        search_in_schedule = ((search_date - first_date).days // 7) + 1  # 计算处在第几周
        # 格式化日期：strftime('%w'),  %w 星期, 0-6 (0是星期天)
        search_week = week_list[int(search_date.strftime('%w'))]  # 计算处在星期几
        class_table_msg = "【课表提醒】\n" + Curriculum.excel_to_print(which_day, raw_message.split(' ')[2], search_in_schedule,search_week, search_date)
    except FileNotFoundError as fe:
        return f"{raw_message.split(' ')[2]} 的课表目前还没有收录到数据库"
    except Exception as e:
        return "输入格式有误\n" \
               "例：课表 后天 19物一  \n"
    else:
        return class_table_msg


# 定义一个课程类
class Curriculum:
    # 星期几对应excel表中的哪一列
    week_to_char = {
        '星期一': 'B',
        '星期二': 'C',
        '星期三': 'D',
        '星期四': 'E',
        '星期五': 'F',
        '星期六': 'G',
        '星期日': 'H'
    }
    # 把excel中的行转换为对应的上课时间
    row_to_jie = {
        4: '时间:8:00-9:40',
        5: '时间:10:00-11:40',
        6: '时间:14:00-15:40',
        7: '时间:16:00-17:40',
        8: '时间:19:00-20:40'
    }

    def __init__(self, class_time, name, class_classroom):
        self.class_time = class_time  # 上课时间
        self.name = name  # 课程名称
        self.class_classroom = class_classroom  # 上课地点

    # 输入参数，schedule: 第几周， week： 星期几， day： 几月几日
    # which_day: 今天、明天、后天
    # class_tag: 19物一
    @classmethod
    def excel_to_print(cls, which_day, class_tag, schedule, week, day):
        """
        第二步: 从文件中读取课表
        """
        wb = load_workbook(f'../static/{class_tag}.xlsx')  # 读取excel档案
        ws = wb.active  # 打开预设的工作表,等价于 -> ws = wb['Sheet1']
        class_num = 0  # 用来统计今天的课程数
        curriculum_list = []  # 用来储存curriculum对象
        for row in range(4, 8 + 1):  # 读取第4行到第8行,即第一二节、第三四节、第五六节、第七八节、第九十节
            char = Curriculum.week_to_char[week]  # 读取第char列的数据
            my_class = ws[char + str(row)].value  # 按char列读取当天的课表
            if my_class is not None:
                start_week = int("".join(re.findall(r"\n(\d+)", my_class)))  # 用正则表达式匹配课程的起始周， 4-15([周])， 4前面是换行\n，4后面是-
                end_week = int("".join(re.findall(r"(\d+)\(", my_class)))  # 用正则表达式匹配课程的结束周， 4-15([周])， 15前面是-， 15后面是(
                if start_week <= schedule <= end_week:  # 如果说当前schedule处在课程教学周，代表今天有这节课
                    class_num = class_num + 1
                    class_name = "".join(re.findall(r"^(.*)", my_class))
                    classroom = "".join(re.findall(r"]\n(.*)$", my_class))
                    curriculum_list.append(
                        Curriculum(Curriculum.row_to_jie[row], class_name, classroom))  # 加入到今天的课程list中

        class_table_msg = ""
        if class_num != 0:
            class_table_msg = class_table_msg + f"{class_tag}的同学们，你们{which_day}有{class_num}节课:\n\n" \
                                                "课表信息:\n"
            for class_info in curriculum_list:
                class_table_msg = class_table_msg + f"{class_info.name}\n" \
                                                    f"{class_info.class_time}\n" \
                                                    f"地点:{class_info.class_classroom}\n\n"
            class_table_msg = class_table_msg + f"上课日期:\n" \
                                                f"第{schedule}周 {week}({day.strftime('%m月%d日')}）"
        else:
            class_table_msg = class_table_msg + f'{which_day}没课，好好happy吧！'
        return class_table_msg


if __name__ == '__main__':
    data = send_class_table("课表 今天 21数二")
    print(data)
